from openpyxl import load_workbook
from tempfile import NamedTemporaryFile


import os
from dotenv import load_dotenv, find_dotenv
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State,StatesGroup
from aiogram import types,Dispatcher
from create_bot_doshkol import dp,bot

from keyboards.admin_Inline_kb import admin_panel,schools,time_and_days,back_kb,is_right,kb_for_edit_child,edit_kb
from aiogram.types import InlineKeyboardMarkup,InlineKeyboardButton,ReplyKeyboardRemove,callback_query,ChatActions,ContentType
from aiogram.dispatcher.filters import Text, ContentTypeFilter

load_dotenv(find_dotenv())
excel_path=os.getenv('excel_path')

class FsmAdmin1(StatesGroup):
    ready_to_fill=State()

print("exel is connected")



async def insert_child(state):
    arr_to_insert=[]
    async with state.proxy() as data:
        workbook = load_workbook(excel_path)
        worksheet = workbook['data']

        async def get_empty_row():
            max_row = worksheet.max_row
            for i in range(max_row, 0, -1):
                if not worksheet[f'B{i}'].value:
                    return i
            return max_row + 1

        row_to_insert = await get_empty_row()
        worksheet[f'A{row_to_insert}'].value = row_to_insert
        worksheet[f'B{row_to_insert}'].value = data['fio']
        worksheet[f'C{row_to_insert}'].value = data['school']
        worksheet[f'D{row_to_insert}'].value = data['time']
        worksheet[f'E{row_to_insert}'].value = data['fioparents']
        worksheet[f'F{row_to_insert}'].value = data['number']
        worksheet[f'G{row_to_insert}'].value = data['agrement']



        workbook.save(excel_path)
        workbook.close()


# async def change_path(new_path):
#     congifg.excel_path=new_path
#     return


async def read_child(school,time):
    workbook = load_workbook(excel_path)
    worksheet = workbook['data']
    print('[')
    child_arr=[]

    for i in range (worksheet.max_row,0,-1):

        if worksheet[f'B{i}'].value:
            print('найдена запись')
            if worksheet[f'C{i}'].value==school and worksheet[f'D{i}'].value==time:
                print('запись передана боту')
                child_arr.append(worksheet[f'B{i}'].value)


    workbook.save(excel_path)
    workbook.close()

    return child_arr

async def info_child(child):
    workbook=load_workbook(excel_path)
    worksheet =workbook['data']
    print("поиск информации о ребёнке")

    info_child_dict={}

    # for i in range(worksheet.max_row, 0, -1):
    #     if worksheet[f'B{i}'].value:
    #         if child== worksheet[f'B{i}'].value:
    #             info_child_dict['fio']=worksheet[f'B{i}'].value
    #             info_child_dict['school'] = worksheet[f'C{i}'].value
    #             info_child_dict['time'] = worksheet[f'D{i}'].value
    #             info_child_dict['fio_parents'] = worksheet[f'E{i}'].value
    #             info_child_dict['number'] = worksheet[f'F{i}'].value
    for i in range(worksheet.max_row,0,-1):
        if worksheet[f'B{i}'].value:
            if child == worksheet[f'B{i}'].value:
                info_child_dict['fio']=worksheet[f'B{i}'].value
                info_child_dict['school'] = worksheet[f'C{i}'].value
                info_child_dict['time'] = worksheet[f'D{i}'].value
                info_child_dict['fioparents'] = worksheet[f'E{i}'].value
                info_child_dict['number'] = worksheet[f'F{i}'].value
                info_child_dict['agrement']=worksheet[f'G{i}'].value


    return info_child_dict
#дописать функцию удаления
async def delete_from_xl(child_dict):
    workbook=load_workbook(excel_path)
    worksheet=workbook['data']
    print('удаление данных о ребёнке')
    print(child_dict)
    for i in range(worksheet.max_row, 0, -1):

            if child_dict['fio']==worksheet[f'B{i}'].value and child_dict['school']==worksheet[f'C{i}'].value and child_dict['time']==worksheet[f'D{i}'].value and child_dict['fioparents']==worksheet[f'E{i}'].value and child_dict['number']==worksheet[f'F{i}'].value:

                worksheet[f'B{i}'].value = ''
                worksheet[f'C{i}'].value = ''
                worksheet[f'D{i}'].value = ''
                worksheet[f'E{i}'].value = ''
                worksheet[f'F{i}'].value = ''
                worksheet[f'G{i}'].value = ''


                for row in range(i + 1, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row - 1, column=col, value=worksheet.cell(row=row, column=col).value)


                for col in worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row):
                    for cell in col:
                        cell.value = None

    workbook.save(excel_path)
    workbook.close()

async def edit_param(callback_query, child_dict, param, state):
    workbook = load_workbook(excel_path)
    worksheet = workbook['data']
    print("перезапись")
    for i in range(worksheet.max_row, 0, -1):
        if worksheet[f'B{i}'].value:
            if child_dict['fio'] == worksheet[f'B{i}'].value and child_dict['school'] == worksheet[f'C{i}'].value and child_dict['time'] == worksheet[f'D{i}'].value and child_dict['fioparents'] == worksheet[f'E{i}'].value and child_dict['number'] == worksheet[f'F{i}'].value:
                print(worksheet[f'B{i}'])
                await state.update_data(param=param)  # Сохранение параметра в состоянии FSM
                await state.update_data(i=i)  # Сохранение i в состоянии FSM



async def edit_in_xl(callback_query,state):
    async with state.proxy() as data:
        param=data['param']

    print("функция", param)
    params={"fio":"ФИО","school":"филлиал","time":"время","fioparents":"ФИО родителей","number":"номер телефона"}


    if param == 'school':
            await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                                  message_id=callback_query.message.message_id,
                                                  text=f'Введите новый {params[param]}', reply_markup=schools,
                                                  parse_mode='html')
    elif param == 'time':
            print('функция времени бля')
            await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                                  message_id=callback_query.message.message_id,
                                                  text=f'Введите новое {params[param]}', reply_markup=time_and_days,
                                                  parse_mode='html')
    elif param == 'number':
            await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                                  message_id=callback_query.message.message_id,
                                                  text=f'Введите номер в нужном формате : +7**********', reply_markup=None,
                                                  parse_mode='html')
    elif param =='back':
        pass

    else:
            await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                          message_id=callback_query.message.message_id,
                                          text=f'Введите новое {params[param]}', reply_markup=None,
                                          parse_mode='html')
    print('вызодит парам'+param)
    return param


async def fill_new_param(message: types.Message, state: FSMContext):
    print("ФУНКЦИЯ ЗАПУСТИЛАСЬ")
    async with state.proxy() as data:
        param=data['param']
        i=data['i']

    print(param,i)
    print("test 1")
    # async with state.proxy() as data:
    #     data[param] = message.text
    workbook = load_workbook(excel_path)
    worksheet = workbook['data']
    print('test 2')
    if param == 'fio':
        worksheet[f'B{i}'] = message.text
    if param == 'school':
        worksheet[f'C{i}'] = message.text
    if param == 'time':
        worksheet[f'D{i}'] = message.text
    if param == 'fioparents':
        worksheet[f'E{i}'] = message.text
    if param == 'number':
        worksheet[f'F{i}'] = message.text

    print('test 3')



    workbook.save(excel_path)


async def fill_new_param1(callback_query:types.CallbackQuery, state: FSMContext):
    print("ФУНКЦИЯ ЗАПУСТИЛАСЬ")
    async with state.proxy() as data:
        param = data['param']
        i = data['i']

    print(param, i)
    print("test 1")
    # async with state.proxy() as data:
    #     data[param] = message.text
    workbook = load_workbook(excel_path)
    worksheet = workbook['data']
    print('test 2')

    if param == 'school':
        worksheet[f'C{i}'] = callback_query.data.split('_')[1]
    if param == 'time':
        worksheet[f'D{i}'] = callback_query.data.split('_')[1]
    if param == 'number':
        worksheet[f'F{i}'] = callback_query.data.split('_')[1]

    print('test 3')

    workbook.save(excel_path)


async def save_xl_file():
    workbook = load_workbook(excel_path)
    temp_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
    workbook.save(temp_file.name)
    workbook.close()
    return temp_file

# @dp.message_handler(state=FsmAdmin1.ready_to_fill)
# async def fill_new_param(message: types.Message, state: FSMContext):
#     print("ФУНКЦИЯ ЗАПУСТИЛАСЬ")
#     async with state.proxy() as data:
#         param=data['param']
#         i=data['i']
#
#     print(param,i)
#     print("test 1")
#     # async with state.proxy() as data:
#     #     data[param] = message.text
#     workbook = load_workbook(excel_path)
#     worksheet = workbook['data']
#     print('test 2')
#     if param == 'fio':
#         worksheet[f'B{i}'] = message.text
#     if param == 'school':
#         worksheet[f'C{i}'] =message.text
#     if param == 'time':
#         worksheet[f'D{i}'] = message.text
#     if param == 'fio_parents':
#         worksheet[f'E{i}'] = message.text
#     if param == 'number':
#         worksheet[f'F{i}'] = message.text
#
#     print('test 3')
#
#     print('test 4')
#
#     workbook.save(excel_path)


# async def edit_param(callback_query,child_dict,param,state):
#     workbook = load_workbook(excel_path)
#     worksheet = workbook['data']
#     print("перезапись")
#     for i in range(1,worksheet.max_row):
#         if worksheet[f'B{i}'].value:
#             if child_dict['fio'] == worksheet[f'B{i}'].value and child_dict['school'] == worksheet[f'C{i}'].value and child_dict['time'] == worksheet[f'D{i}'].value and child_dict['fio_parents'] == worksheet[f'E{i}'].value and child_dict['number'] == worksheet[f'F{i}'].value:
#                     print(worksheet[f'B{i}'])
#                     async def edit_in_xl(callback_query,param):
#                         print("функция ",param)
#                         if param=='fio':
#                             await bot.edit_message_text(chat_id=callback_query.message.chat.id,
#                                                         message_id=callback_query.message.message_id,
#                                                         text=f'Введите новое фио',
#                                                         parse_mode='html')
#
#
#                         if param=='school':
#                             await bot.edit_message_text(chat_id=callback_query.message.chat.id,
#                                                         message_id=callback_query.message.message_id,
#                                                         text=f'Введите новую школу',
#                                                         parse_mode='html')
#
#                         if param=='time':
#                             await bot.edit_message_text(chat_id=callback_query.message.chat.id,
#                                                         message_id=callback_query.message.message_id,
#                                                         text=f'Введите новое время',
#                                                         parse_mode='html')
#
#                         if param=='fio_parents':
#                             await bot.edit_message_text(chat_id=callback_query.message.chat.id,
#                                                         message_id=callback_query.message.message_id,
#                                                         text=f'Введите новое фио родителя',
#                                                         parse_mode='html')
#
#                         if param=='number':
#                             await bot.edit_message_text(chat_id=callback_query.message.chat.id,
#                                                         message_id=callback_query.message.message_id,
#                                                         text=f'Введите новый номер телефона',
#                                                         parse_mode='html')
#                         print("значение param перед входом в функцию:"+param)
#                         await edit_in_xl(callback_query,param)
#                         print('[qq')
#                         await fill_new_param(param, callback_query.message, state)
#
#                     @dp.message_handler()
#                     async def fill_new_param(param,message:types.Message,state:FSMContext):
#                         print('бля эта функция сработала')
#                         # async with state.proxy() as data:
#                         #     data[str(param)]=message.text
#                         if param=='fio':
#                             worksheet[f'B{i}'] = message.text
#                             return
#                         if param=='school':
#                             worksheet[f'C{i}'] = message.text
#                             return
#                         if param=='time':
#                             worksheet[f'D{i}'] = message.text
#                             return
#                         if param=='fio_parents':
#                             worksheet[f'E{i}'] = message.text
#                             return
#                         if param=='number':
#                             worksheet[f'F{i}'] = message.text
#                             return
#
#
#
#
#     workbook.save(excel_path)
#     workbook.close()
